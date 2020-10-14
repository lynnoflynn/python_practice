from openpyxl import Workbook, load_workbook
#定义一个类
class PracticeExcel:
    def create(self):
        wb = Workbook()
        ws = wb.active
        ws['A1'] = "身高"
        ws['B1'] = "体重"
        self.height = [180, 160, 170, 155]
        weight = [66, 50, 40, 30]
        for i in range(len(self.height)):
            # 填入身高信息
            ws.cell(row=i + 2, column=1, value=self.height[i])
            # 填入体重信息
            ws.cell(row=i + 2, column=2, value=weight[i])
        # save the file
        wb.save("sample.xlsx")
    def health_weight(self):
        #找到excel文档和页签
        ld = load_workbook(filename="sample.xlsx", )
        sheet = ld.active
        sheet["C1"] = "备注"
        for i in range(len(self.height)):
            height = sheet.cell(row=i + 2, column=1).value
            weight = sheet.cell(row=i + 2, column=2).value
        #获取健康体重
            health_weight = (height-70)*0.6
            if weight == health_weight:
                print("这是健康体重哦~",weight)
                sheet.cell(row=i + 2, column=3).value="健康体重"
        # save the file
        ld.save("sample1.xlsx")
pe = PracticeExcel()
pe.create()
pe.health_weight()