from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws['A1'] = "身高"
ws['B1'] = "体重"
height = [160,166,158]
weight = [50,58,45]
for i in range (len(height)):
    #填入身高信息
    ws.cell(row = i+2,column=1, value=height[i])
    #填入体重信息
    ws.cell(row = i+2,column=2, value=weight[i])
# save the file
wb.save("sample.xlsx")